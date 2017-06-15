# -*- coding: utf-8 -*-
import sys
import os
import csv
import xlrd
import shutil
from xlutils.copy import copy
from openpyxl import load_workbook

global gl_sheet_name
global gl_excel_path
global is_lower
global sheet_index

def write_excel():
    global gl_excel_path,gl_sheet_name,is_lower
    if is_lower:
        gl_sheet_name = gl_sheet_name.name
    if gl_excel_path == None or gl_excel_path == '':
        print('None gl path')
        return
    
    data = xlrd.open_workbook(gl_excel_path)
    file_name = gl_excel_path.split('\\')[-1]
    tmpData = copy(data)
    csvfile = open(gl_sheet_name+"_2.csv","rb")
    from csv import reader
    readers = reader(csvfile)
    n = 0
    m = 0
    for row in readers:
        for i in row:
            tmpData.get_sheet(gl_sheet_name).write(n,m,unicode(i,'utf-8'))
            m = m + 1
        else:
            m = 0
            n = n +1
    os.remove(gl_excel_path)
    tmpData.save(file_name)

def write_excel_higher():
    global gl_excel_path, gl_sheet_name, is_lower
    if is_lower:
        gl_sheet_name = gl_sheet_name.name
    if gl_excel_path == None or gl_excel_path == '':
        print('None gl path')
        return
    file_name = gl_excel_path.split("\\")[-1]
    shutil.move(gl_excel_path,file_name)

    data = load_workbook(file_name)
    global sheet_index
    tmpData = data.worksheets[int(sheet_index)]
    csvfile = open(gl_sheet_name + "_2.csv", "rb")
    from csv import reader
    readers = reader(csvfile)
    n = 1
    m = 1
    for row in readers:
        for i in row:
            tmpData.cell(row= n,column = m,value = unicode(i,'utf-8'))
            m = m + 1
        else:
            m = 1
            n = n + 1
    os.remove(file_name)
    data.save(file_name)

def eachFile(filepath):
    pathDir =  os.listdir(filepath)
    res = []
    for allDir in pathDir:
        child = os.path.join('%s\%s' % (filepath, allDir))
        if child.split('.')[-1] == 'xls' or child.split('.')[-1] == 'xlsx':
            res.append(child.decode('gbk'))
    else:
        return res

def get_excel_name():
    Excel_1 = ''
    Excel_2 = ''

    res = eachFile('Excel_1')
    for file in res:
        Excel_1 = file
    res = eachFile('Excel_2')
    for file in res:
        Excel_2 = file

    excel2csv(Excel_1,Excel_2)
    return Excel_1,Excel_2

def create_csv(filename,sheet):
    if sheet == None:
        print("no such sheet!!")
        return
    try:
        num = filename[6:7]
        xlsx_file_reader = load_workbook(filename)
        csv_filename = sheet.encode('utf_8')+'_'+str(num)+'.csv'
        csv_file = file(unicode(csv_filename,'utf-8'), 'wb')
        csv_file_writer = csv.writer(csv_file)
        sheet_ranges = xlsx_file_reader[sheet]

        for row in sheet_ranges.rows:
            row_container = []
            for cell in row:
                if type(cell.value) == unicode:
                    value = cell.value.encode('utf-8')
                    if value == 'None':
                        value = ''
                    row_container.append(value)
                else:
                    value = str(cell.value)
                    if value == 'None':
                        value = ''
                    row_container.append(value)
            csv_file_writer.writerow(row_container)
        csv_file.close()
    except Exception as e:
        print e

def create_csv_lower(filename,sheet):
    if sheet == None:
        print("no such sheet!!")
        return
    try:
        num = filename[6:7]
        xlsx_file_reader = xlrd.open_workbook(filename)
        csv_filename = sheet.name.encode('utf_8')+'_'+str(num)+'.csv'
        csv_file = file(unicode(csv_filename,'utf-8'), 'wb')
        csv_file_writer = csv.writer(csv_file)

        for row in range(sheet.nrows):
            row_container = []
            for cell in sheet.row_values(row):
                if type(cell) == unicode:
                    value = cell.encode('utf-8')
                    if value == 'None':
                        value = ''
                    row_container.append(value)
                else:
                    value = str(cell)
                    if value == 'None':
                        value = ''
                    row_container.append(value)
            csv_file_writer.writerow(row_container)
        csv_file.close()
    except Exception as e:
        print e


def excel2csv(filename,filename2):
    csv_filename = ''
    csv_filename1 = ''
    if filename == '':
        print 'import excel is None!'
        sys.exit(0)
    elif (filename.split('.')[-1]).endswith('x'):
        print "office 2007 and a higher version"
        global is_lower
        is_lower = False;
        try:
            xlsx_file_reader = load_workbook(filename)
            sheets = []
            for sheet in xlsx_file_reader.get_sheet_names():
                sheets.append(sheet)
            i = 0
            for sheet in sheets:
                print(str(i)+":"+sheet)
                i = i +1;
            index = raw_input("choose sheet:")
            global sheet_index
            sheet_index = index
            global gl_sheet_name
            choose_sheet = sheets[int(index)]
            gl_sheet_name = choose_sheet;
            create_csv(filename,choose_sheet)
            create_csv(filename2,choose_sheet)
            global  gl_excel_path
            gl_excel_path = filename2
        except Exception as e:
            print e
        return csv_filename
    elif (filename.split('.')[-1]).endswith('s'):
        print "office 2007 and a lower version"
        global is_lower
        is_lower = True;
        try:
            xls_file_reader = xlrd.open_workbook(filename)
            length = len(xls_file_reader.sheets())
            sheets = []
            for sheet in xls_file_reader.sheets():
                sheets.append(sheet)
            i = 0
            for sheet in sheets:
                print(str(i) + ":" + sheet.name)
                i = i + 1;
            index = raw_input("choose sheet:")
            global sheet_index
            sheet_index = index
            global gl_sheet_name
            choose_sheet = sheets[int(index)]
            gl_sheet_name = choose_sheet;
            create_csv_lower(filename, choose_sheet)

            xls_file_reader2 = xlrd.open_workbook(filename2)
            sheets2 = []
            for sheet2 in xls_file_reader2.sheets():
                sheets2.append(sheet2)
            choose_sheet2 = sheets2[int(index)]
            create_csv_lower(filename2, choose_sheet2)

            global gl_excel_path
            gl_excel_path = filename2

        except Exception as e:
            print e
        return csv_filename1

if not os.path.exists('Excel_1'):
    os.mkdir('Excel_1')

if not os.path.exists('Excel_2'):
    os.mkdir('Excel_2')

get_excel_name()

do = raw_input("create_Excel? y/n")
if do == 'y':
    if is_lower:
        write_excel()
    else:
        write_excel_higher()
else:
    print("baibai")
    os._exit(0);